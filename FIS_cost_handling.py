# ---
# jupyter:
#   jupytext:
#     formats: ipynb,py:light
#     text_representation:
#       extension: .py
#       format_name: light
#       format_version: '1.5'
#       jupytext_version: 1.15.0
#   kernelspec:
#     display_name: Python 3 (ipykernel)
#     language: python
#     name: python3
# ---

import pandas as pd
import time
import os
from datetime import datetime
import xlrd
from tkinter import filedialog as fd
from tkinter import messagebox as mb
import tkinter as tk
import win32com.client as win32
import shutil
import re
import openpyxl as op
from openpyxl import load_workbook
import numpy as np
from openpyxl.utils.dataframe import dataframe_to_rows
from tkinter import *
from tkinter import font
from tkinter import ttk
import math
import numpy as np


# +
def func(event):
    #print("First Input is ",tk.Entry.get(textBox))
    
    root.destroy()
    

def show_selected():
    root.destroy()  # Close the GUI window after getting input

# +
#ext_initial_width = 300 ; text_initial_height = 150


text_initial_width = 400 ; text_initial_height = 150

root=tk.Tk()
root.title("파일명 Input #1 - 업체명")
root.geometry(f"{text_initial_width}x{text_initial_height}")


#def retrieve_input():
#    input_value=textBox.get("1.0","end-1c")
#    print(input_value)
#    return input_value"

var_str = StringVar()
input_font = tk.font.Font(size = 13, weight = "bold")

textBox=ttk.Entry(root, width=20, textvariable = var_str,font = input_font )
textBox.pack(pady=20)

enter_font =  font.Font(size=11, weight = "bold")
root.bind('<Return>', func)


buttonCommit=Button(root, height=1, width=10, text="Enter", fg = "blue", font = enter_font, 
                    command= show_selected)
#command=lambda: retrieve_input() >>> just means do this when i press the button
buttonCommit.pack(pady=20)

#show_button = tk.Button(root, text="Enter", command=show_selected, fg="blue")
#show_button.pack()

mainloop()

first_input = var_str.get()
print(first_input)


# Second Input

#ext_initial_width = 300 ; text_initial_height = 150
root=tk.Tk()
root.title("파일명 Input #2-운송방법 ")
text_inital_width = 400 ; text_initial_height = 150
root.geometry(f"{text_initial_width}x{text_initial_height}")


#def retrieve_input():
#    input_value=textBox.get("1.0","end-1c")
#    print(input_value)
#    return input_value"

var_str = StringVar()
input_font = tk.font.Font(size = 13, weight = "bold")

textBox=ttk.Entry(root, width=25, textvariable = var_str,font = input_font )
textBox.pack(pady=20)

enter_font =  font.Font(size=11, weight = "bold")
root.bind('<Return>', func)


buttonCommit=Button(root, height=1, width=10, text="Enter", fg = "blue", font = enter_font, 
                    command= show_selected)
#command=lambda: retrieve_input() >>> just means do this when i press the button
buttonCommit.pack(pady=20)

#show_button = tk.Button(root, text="Enter", command=show_selected, fg="blue")
#show_button.pack()

mainloop()

second_input = var_str.get()
print(second_input)


input_name = first_input + "_" + second_input + "_" + "Total_"

final_input_name = ""

try: 
    temp = input_name.split(" ")
    filter_input_name = list(filter(None, temp))

    for i in range(len(filter_input_name)):
    # print(i)
        final_input_name += filter_input_name[i]

except:
    
    final_input_name = input_name
    

# +

#print(final_input_name)


def display_input():
    input_value = []
    if var1.get() == 1:
        input_value = "영동 글로벌 / ACE"
        print("1. Input for option 1:", var1.get(), input_value)
    else:
        input_value = "영세 / 과세 / Wharfage"
        print("1. Input for option 2:", var2.get(), input_value)

    return input_value



def Xlsx_file():
    root = tk.Tk()
    file = fd.askopenfilenames(parent=root, title="""정산내역 파일을 선택해주세요""",
                               filetypes=[("Excel files", "*xlsx")])  # seems XLSX works!

    if file == "":
        mb.showwarning("Warning", "정산 파일 선택해주세요. (.xlsx 확장자)")
    root.destroy()

    # print(file)
    return file

def Xls_file():
    root = tk.Tk()
    cost_center_files = fd.askopenfilenames(parent=root, title="""Cost Center 파일 선택해주세요. (.xls 확장자) """,
                                            filetypes=(("xls", "*xls"), ("xlsx", "*xlsx")))

    if cost_center_files == "":
        mb.showwarning("Warning", "Cost Center 파일 선택해주세요. (확장자 구분 필요)")

    root.destroy()

    # print(cost_center_files)

    return cost_center_files



# # +
# 파일 변환 - 공통

def xls_to_xlsx(cost_center_files,xlsx_new_folder):
    xlsx_fullname_list = []

    for i in range(len(cost_center_files)):
        file_name = cost_center_files[i].split("/")[-1]

        excel = win32.gencache.EnsureDispatch('Excel.Application')

        try:
            wb = excel.Workbooks.Open(os.path.abspath(cost_center_files[i]))
            wb.SaveAs(os.path.abspath(cost_center_files[i]) + "X", FileFormat=51)
            wb.Close()
            print("    - {} out of {} : File Well Created!".format(i+1, len(cost_center_files)))

        except:
            print("Error Occured!")

        temp = xlsx_new_folder + "/" + file_name + "X"
        shutil.move(cost_center_files[i] + "X", temp)

        # temp = xlsx_new_folder + "/" + file_name + "X"
        xlsx_fullname_list.append(temp)
        temp = ""

    excel.Application.Quit()

    # print(xlsx_fullname_list)

    print("3. SAP Files convert done (XLS --> XLSX!) ")

    return xlsx_fullname_list


# 오직 영동글로벌만 !!!!

def sap_file_merge_into_one(xlsx_new_folder, ori_c_filename):
    sheetname = ori_c_filename.split("_")[0]
    ends_with = ['.XLSX', '.xlsx']

    excl_list = []

    for filename in os.listdir(xlsx_new_folder):

        if filename.endswith(
                tuple(ends_with)):  # file names end with "xls" & "xlsx" / endswith accepts tuple data types
            file_path = os.path.join(xlsx_new_folder, filename)
            # print(file_path)

            excl_list.append(pd.read_excel(file_path))  # DataFrame append to list

    excl_merged = pd.concat(excl_list, ignore_index=True)
    excl_output = xlsx_new_folder + "/" + sheetname + "_" + output_folder + "_merged.xlsx";
    excl_output
    # excl_merged.to_excel(excl_output,sheet_name = sheetname, index = False )

    return excl_merged, excl_output, sheetname


# excl_merged = merged_df
# excl_ouptut  = merged 이름 (but 엑셀로 빼지는 않음)
# sheetname (ex) 230520A3


# 공통
# 정규표현식 사용하여 해당 단어 포함하고 있는 열 인덱스 찾아내기

def Matching_words(col_list, partial_list):
    matching_words = []
    matching_col_index = []

    for i in range(len(partial_list)):

        partial = partial_list[i]
        pattern = re.compile(rf'\w*{partial}\w*')

        num = 0

        for word in col_list:
            if re.match(pattern, str(word)):
                matching_words.append(word)
                matching_col_index.append(num)

            num += 1

    #print(matching_words, matching_col_index)

    return matching_words, matching_col_index

# Create list of lists --> Nested List

def Create_list(n):
    list_of_list = []

    for _ in range(n):
        list_of_list.append([])
    # print(list_of_lists)

    return list_of_list

# 공통
# SAP 파일 BL 공백 없애기 

def sap_strip_BL(second_index):
    for i in range(len(second_index)):
        a = merged_del_df.iloc[:, second_index[i]]
        # print("***", second_word[i])

        for ii in range(len(a)):
            if type(a[ii]) != type("ABC"):
                #print(a[ii], len(str(a[ii])))

                list_of_list[i].append(a[ii])
            else:
                #print(a[ii].strip(), len(a[ii].strip()))
                list_of_list[i].append(a[ii].strip())
    
    return list_of_list


# common

def common_cleansing(merged_df):
    
    partial_list = ["Order", "WBS"]  # delete col
    second_cols_list = ["Bill", "Account", "Cost", "Prctr"] # preprocessed col

# 1. Delete Col for 2 columns : "Order", "WBS"
    merged_temp_col = merged_df.columns 

    delete_col, delete_index = Matching_words(merged_temp_col, partial_list)
    merged_del_df = merged_df.drop(merged_df.columns[delete_index], axis=1)
# merged_del_df : dataframes w.o order & wbs


# 2. Data cleansing for 4 columns : "Bill", "Account", "Cost", "Prctr"

    col_2 = merged_del_df.columns
    second_word, second_index = Matching_words(col_2, second_cols_list)

    return merged_del_df, second_index, second_word

# 공통 

def sap_processed_four_cols(second_word, list_of_list):
    dataframes_list = []

    # Convert each inner list to a DataFrame with one column and append it to the list
    num = 0
    for inner_list in list_of_list:
        df = pd.DataFrame({second_word[num]: inner_list})
        dataframes_list.append(df)
        num += 1

    # Access the individual DataFrames using the index
    df1 = dataframes_list[0]
    df2 = dataframes_list[1]
    df3 = dataframes_list[2]
    df4 = dataframes_list[3]

    return df1, df2, df3, df4


#  Processed cleansed 4 columns convert to DATAFRAME

# 오직 포워더


def option_2_sheetname(ori_c_filename):
    sheetname = ori_c_filename.split(".")[0]
    
    return sheetname


# 공통

def sap_final_merged(merged_del_df, second_index, z, sheetname, final_input_name):

    merged_col = merged_del_df.columns  # Total columns
    start_num = second_index[0]  ; end_num = second_index[-1]  # index 만들기

    start = [] ; end = []

    for i in range(start_num):
        start.append(i)

    for i in range(end_num+1,len(merged_col)):
        end.append(i)

    start_part = merged_del_df.iloc[:,start] # SAP first col ~ fourth
    end_part = merged_del_df.iloc[:,end]  # last col of SAP : Amount


    part_1 = pd.merge(df1, df2,left_index= True, right_index = True, how = 'left')
    part_2 = pd.merge(part_1, df3,left_index= True, right_index = True, how = 'left')
    middle_part = pd.merge(part_2, df4,left_index= True, right_index = True, how = 'left')
    temp_merged = pd.merge(start_part, middle_part, left_index= True, right_index = True, how = 'left')
    final_merged = pd.merge(temp_merged, end_part, left_index= True, right_index = True, how = 'left') ; final_merged

    
    if var1_output == 1:
        
    #파이널 파일명
        #final_excl = xlsx_new_folder + "/" + "Final_" +sheetname + "_merged.xlsx"
        final_excl = xlsx_new_folder + "/" + final_input_name +sheetname + ".xlsx"

    # 파일널 파잃명 to excel
        final_merged.to_excel(xlsx_new_folder + "/" + final_input_name +sheetname + ".xlsx", index = False, sheet_name = sheetname )
    
    else:
        sheetname = option_2_sheetname(xlsx_fullname_list[z][2:].split("/")[-1])
        final_excl = xlsx_new_folder + "/" + final_input_name + sheetname + ".xlsx"
        final_merged.to_excel(xlsx_new_folder + "/" +final_input_name +  sheetname + ".xlsx", index = False, sheet_name = sheetname )
        
    print("6. Final File Created!")
    
    return final_excl, final_merged, end_part
    
# final_merged = df
# final_excl = 엑셀 경로

# 공통 
def create_sheet(final_excl):
    
    wb = load_workbook(final_excl)
    ws = wb.active
    wb.create_sheet('Sheet1', 1)

    wb.save(final_excl)  # SAP 전처리 완료된 통합된 파일 시트와 sheet 1 새롭게 생성

    print("7. Sheet1 in Final File has been created")
    
    
def read_xlsx(file):
    try:
        xlsx_df = pd.read_excel(file[0], sheet_name = "SAP Upload")

    except PermissionError:
        print("**** ERROR : Please close the excel file.")
        
    if var1_output ==1: # 영동 & ACE 라면
        copy_col = ["BL", "Posting Amt"]
        
    else:
        copy_col = ["BL", "Posting Amt", "Index"]
    
    
    xlsx_col = xlsx_df.columns.tolist() 
    xlsx_col, xlsx_index = Matching_words(xlsx_col, copy_col)

    copy_excl_temp = xlsx_df.iloc[:, xlsx_index]

    return xlsx_col, xlsx_index, copy_excl_temp, xlsx_df



def main_strip_xlsx(xlsx_index, master_excl_df):

    if var1_output == 1:
        copy_list_of_list = Create_list(2)
    else:
        copy_list_of_list = Create_list(3)

    for i in range(len(xlsx_index)):
        a = master_excl_df.iloc[:, i]    # 기존 전체 df에서 2,5 인덱스 뽑는 것이 아닌, 0,1로 변경
    #print(len(a))
        for ii in range(len(a)):
        #print(a[ii], type(a[ii]), len(a[ii]))
        #print(a[ii], type(a[ii]))
             # only for BL
            if i == 0: # BL 이라면 
                try:
                    strip_data = int(a[ii])
                #print(i, "str", strip_data, type(strip_data))
                    
                except:
                    temp = str(a[ii])
                    strip_data = temp.strip()
            #    print(i,"int", strip_data, type(strip_data))
                
            #print(strip_data, len(str(strip_data)), ii) 
            
                copy_list_of_list[i].append(strip_data)
        
            else:  # numbers - round
            
                round_num = round(a[ii])
                #print(i, round_num, ii)
                copy_list_of_list[i].append(round_num)
        #    


    print("8. data preprocessing for 정산 파일 is done")
    
    return copy_list_of_list


    
def main_xlsx_dataframe_var1(copy_list_of_list, xlsx_col):
    
    dataframes_list = []
    
    num = 0
    for inner_list in copy_list_of_list:
        df = pd.DataFrame({xlsx_col[num]: inner_list})
        dataframes_list.append(df)
        num += 1

    copy_df_1 = dataframes_list[0]
    copy_df_2 = dataframes_list[1]
    
    copy_final_df = pd.merge(copy_df_1, copy_df_2, left_index=True, right_index=True, how='left')

    return copy_final_df


def main_xlsx_dataframe_var2(copy_list_of_list, xlsx_col):
    
    dataframes_list = []
    
    num = 0
    for inner_list in copy_list_of_list:
        df = pd.DataFrame({xlsx_col[num]: inner_list})
        dataframes_list.append(df)
        num += 1

    copy_df_1 = dataframes_list[0]
    copy_df_2 = dataframes_list[1]
    copy_df_3 = dataframes_list[2]
    
        
    temp = pd.merge(copy_df_1, copy_df_2, left_index=True, right_index=True, how='left')
    copy_final_df = pd.merge(temp, copy_df_3, left_index=True, right_index=True, how='left')


    return copy_final_df


def pivot_sap_file(df1, z):
    
    df_1_colname = df1.columns[0] # Bill of Lading

    vl_df1 = df1.rename(columns={df_1_colname:'BL'})
    
    #end part = SAP Amount
    xls_vlookup = pd.merge(vl_df1,end_part, left_index= True, right_index = True, how = 'left') ; xls_vlookup # SAP 합친 파일 Bill of Lading 과 BL 맞추기
    sum_xls_df = xls_vlookup.groupby('BL').sum()
    sum_xls_df.reset_index(inplace = True) # 원본 변경

    
    if var1_output == 1:
        copy_df = pd.merge(copy_final_df, sum_xls_df, how = "left") 
        
    else:
        index_num = copy_final_df["Index"] == z+1  # 여기다가 for range
        index_filter = copy_final_df[index_num].iloc[:,:2]
        copy_df = pd.merge(index_filter, sum_xls_df, how = "left")
        
        
    na_sap_amount = copy_df[copy_df.columns[-1]].tolist() # amount
    master_amount = copy_df[copy_df.columns[1]].tolist()  # posting amt
        
        
    diff = []
    check = []

    for i in range(len(na_sap_amount)):
    
        na_check = math.isnan(na_sap_amount[i])
        temp = master_amount[i] - na_sap_amount[i] 
    
        if na_check == True:
            check.append("")
        else:
            if -20 <= temp <= 20:
                check.append("")
            else: 
                check.append("*")


    #diff.append(temp)
#print(len(diff))


    check_df = pd.DataFrame(check, columns = ["**"])
    temp = pd.merge(copy_df, check_df,left_index= True, right_index = True, how = 'left')
    temp.reset_index(inplace = True)
    final_check_df = temp.iloc[:, 1:]    
        
        
    return copy_df, final_check_df



def final_save(final_excl, copy_df, z):

    book = load_workbook(final_excl)

# Create a Pandas Excel writer using the ExcelWriter class and the loaded workbook
    with pd.ExcelWriter(final_excl, engine='openpyxl') as writer:
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    # Write the DataFrame to the specific sheet named 'Sheet1' starting at row 4 (after existing data)
        copy_df.to_excel(writer, sheet_name='Sheet1', index=False)

    # startrow=writer.sheets['Sheet1'].max_row) # start writing the DataFrame's data immediately after the last row of data
    
    
    if var1_output ==1 :
        print("Complete!","Please check the file ---> ", final_excl)
        
    else:
        print("Complete!",z+1,"번째","Please check the file ---> ", final_excl)
        
    print("-"*100)

# +
def final_save_master(final_excl, final_copy_df, z, sheetname):

    book = load_workbook(final_excl)

# Create a Pandas Excel writer using the ExcelWriter class and the loaded workbook
    with pd.ExcelWriter(final_excl, engine='openpyxl') as writer:
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    # Write the DataFrame to the specific sheet named 'Sheet1' starting at row 4 (after existing data)
        final_copy_df.to_excel(writer, sheet_name=sheetname, index=False)

    # startrow=writer.sheets['Sheet1'].max_row) # start writing the DataFrame's data immediately after the last row of data
    
    
    print("     - NaN value added !")
    #if var1_output ==1 :
    #    print("Complete!","Please check the file ---> ", final_excl)
        
    #else:
    #    print("Complete!",z+1,"번째","Please check the file ---> ", final_excl)
        
#    print("-"*100)

def cost_handling(final_merged_df_temp):
    cost_col_name = final_merged_df_temp.columns
    cost_col, cost_index = Matching_words(final_merged_df_temp.columns, ["Prctr"])

    cost_1 = final_merged_df_temp.iloc[:, cost_index]

    file_input_1 = r'\\vt1.vitesco.com\smt\did99096\04_Plant_SCM\02_Advanced SCM SC\10_RPA\CFT_FIS_RPA\Cost_Center_Master_2023.xlsx'
    file_input_2 = r'\\icfs2001a.vt1.vitesco.com\did99096\04_Plant_SCM\02_Advanced SCM SC\10_RPA\CFT_FIS_RPA\Cost_Center_Master_2023.xlsx'

    try:
        master_cost = pd.read_excel(file_input_1, sheet_name="2023")
    except:
        master_cost = pd.read_excel(file_input_2, sheet_name="2023")

    cost_temp = pd.merge(cost_1, master_cost, how="left")
    cost_final = cost_temp.iloc[:, -1:]
    # cost_final # pivot 완료된 파일

    temp_1 = pd.merge(final_merged_df_temp.iloc[:, :cost_index[0] - 1], cost_final, left_index=True, right_index=True,
                      how='left')
    final_merged_df = pd.merge(temp_1, final_merged_df_temp.iloc[:, cost_index[0]:], left_index=True, right_index=True,
                               how='left')

    return final_merged_df


# +
root = tk.Tk()
root.title("Type Checking")

initial_width = 300 ; initial_height = 180
root.geometry(f"{initial_width}x{initial_height}")

custom_font = font.Font(size=11, weight = "bold")
enter_font =  font.Font(size=10, weight = "bold")

var1 = IntVar() ; var2 = IntVar()

# option 1
checkbox = tk.Checkbutton(root, text="1. 영동글로벌 / ACE", variable=var1, command = display_input, font = custom_font)
checkbox.pack(anchor="w", pady = 10)  # Align checkboxes to the west (left)

# option 2
checkbox2 = tk.Checkbutton(root, text="2. 영세 / 과세 / Wharfage", variable=var2, command = display_input, font = custom_font)
checkbox2.pack(anchor="w", pady=20)  # Align checkboxes to the west (left)

# Create a button to show selected checkboxes
show_button = tk.Button(root, text="Enter", command=show_selected, fg="blue", font = enter_font )
show_button.pack()


# Label to display selected items
result_label = tk.Label(root, text="")
result_label.pack()

# Start the main loop
root.mainloop()


var1_output = var1.get()
var2_output = var2.get()

#print(var1_output, var2_output)


file = Xlsx_file()
time.sleep(1)
cost_center_files = Xls_file()




# filename -- 파일명 자체 (ex) 230520A3_1_A03.XLS
filename = cost_center_files[0][2:].split("/")[-1]
ori_c_filename = filename

# 파일 전체 길이 - 파일명만 제외
root_path_len = len(cost_center_files[0]) - len(cost_center_files[0][2:].split("/")[-1])

# 파일명 제외한 경로
root_path = cost_center_files[0][:root_path_len]  # 새로운 폴더 생성 직전까지 생김

# newly created folder
output_folder = datetime.now().strftime("Output_%b_%d_%H_%M")  # 'Output_Aug_11_09_18'

# new folder creation
xlsx_new_folder = root_path + output_folder
os.makedirs(xlsx_new_folder)
print("2. xlsx folder created!")


#print(var1_output, var2_output)

time.sleep(2)  # 안그럼 excel pop up 

xlsx_fullname_list = xls_to_xlsx(cost_center_files, xlsx_new_folder)


# +
# excl_merged.to_excel("check_double.xlsx")

# +
def add_na_final(copy_df,final_merged_df):
    
    na_sap_amount = copy_df[copy_df.columns[-1]].tolist() # amount
    master_amount = copy_df[copy_df.columns[1]].tolist()  # posting amt
    
    na_filter = pd.isna(copy_df[copy_df.columns[-1]])
    filter_out = np.where(na_filter == True)
    
    nan_df = copy_df.loc[filter_out]


    nan_add_col = final_merged_df.columns.tolist()

    partial_list = ["Bill of Lading", "Amount"]
    col_name, col_index = Matching_words(nan_add_col, partial_list)

    #col_name, col_index

    nan_df_cols = nan_df.columns

    nan_df_temp = nan_df.rename(columns = {nan_df_cols[0]:col_name[0]})
    nan_df_rename = nan_df_temp.rename(columns = {nan_df_cols[1]:col_name[1]})



    nan_df_rename.reset_index(inplace = True)

    final_nan_df = nan_df_rename.iloc[:,1:3]

#final_nan_df

    final_copy_df = pd.concat([final_merged_df,final_nan_df], ignore_index = True)
    
    
    return final_copy_df


if var1_output == 1 :
    z = ""
    excl_merged, excl_output, sheetname = sap_file_merge_into_one(xlsx_new_folder, ori_c_filename)


    merged_df = excl_merged
    print("*only for 영동/ACE:  SAP file has been merged into one file")
    
    merged_del_df, second_index, second_word = common_cleansing(merged_df)
    
    list_of_list = Create_list(4)
    list_of_list = sap_strip_BL(second_index)
    
    df1, df2, df3, df4 = sap_processed_four_cols(second_word, list_of_list)
    print("5. data preprocessing for SAP 파일 is done (DataFrame)")

    final_excl, final_merged_df_temp, end_part = sap_final_merged(merged_del_df, second_index,z, sheetname, final_input_name)
    final_merged_df = cost_handling(final_merged_df_temp)
    
    create_sheet(final_excl)
    
    xlsx_col, xlsx_index , copy_excl_temp,xlsx_df = read_xlsx(file)
    
    master_excl_df  = copy_excl_temp.groupby('BL').sum()
    master_excl_df.reset_index(inplace = True)
    master_excl_df = master_excl_df.dropna(subset=[xlsx_col[0]])
    
    copy_list_of_list = main_strip_xlsx(xlsx_index,master_excl_df)
    
    copy_final_df = main_xlsx_dataframe_var1(copy_list_of_list, xlsx_col)
    
    copy_df, final_check_df = pivot_sap_file(df1,z) 
    
    final_copy_df = add_na_final(copy_df,final_merged_df)
    
    final_save_master(final_excl, final_copy_df, z, sheetname)
    
    final_save(final_excl, final_check_df, z)
    
    #final_save(final_copy_df, final_check_df, z)
    
    
else:
    
    for z in range(len(xlsx_fullname_list)):
    #print(xlsx_fullname_list[i])
        merged_df = pd.read_excel(xlsx_fullname_list[z])
    
    
        merged_del_df, second_index, second_word = common_cleansing(merged_df)
        
        
        list_of_list = Create_list(4)
        list_of_list = sap_strip_BL(second_index)
        
        df1, df2, df3, df4 = sap_processed_four_cols(second_word, list_of_list)
        print("5. data preprocessing for SAP 파일 is done (DataFrame)")

        sheetname = option_2_sheetname(ori_c_filename)

        final_excl, final_merged_df_temp, end_part = sap_final_merged(merged_del_df, second_index, z, sheetname, final_input_name)
        final_merged_df = cost_handling(final_merged_df_temp)

        create_sheet(final_excl)
        xlsx_col, xlsx_index , copy_excl_temp,xlsx_df = read_xlsx(file)
        
        index_num = list(set(copy_excl_temp[xlsx_col[-1]].tolist()))
    
    
        ex_1 = copy_excl_temp.dropna(subset=[xlsx_col[0]]) # BL 에서 na 행 다 버리기
        filter_1 = ex_1[xlsx_col[-1]]== index_num[z+1]  # index 값

        master_filter = ex_1[filter_1]  # 인덱스 값끼리 데이터 묶기
    
    #각 인덱스로 구분 후 
    
        master_excl_df  = master_filter.groupby('BL').sum()  # 피벗 돌리기 
        master_excl_df.reset_index(inplace = True)
        master_excl_df = master_excl_df.dropna(subset=[xlsx_col[0]])

        
        
        copy_list_of_list = main_strip_xlsx(xlsx_index,master_excl_df)
        
        copy_final_df = main_xlsx_dataframe_var2(copy_list_of_list, xlsx_col)
        
        copy_df,final_check_df = pivot_sap_file(df1, z) 
        
        final_copy_df = add_na_final(copy_df,final_merged_df)
    
        #final_save(final_copy_df, final_check_df, z)
        
        sheetname_ = xlsx_fullname_list[z].split("/")[-1].split(".")[-2]
        #print(sheetname_)
        
        final_save_master(final_excl, final_copy_df, z, sheetname_)
        
        final_save(final_excl, final_check_df, z)

# -
print()
time.sleep(5)
